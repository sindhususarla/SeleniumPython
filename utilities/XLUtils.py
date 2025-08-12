import openpyxl

def getRowCount(file,sheetName):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    return sheet.max_row

def readData(file,sheetName,rowNum, colNum):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    return sheet.cell(row=rowNum, column=colNum).value

